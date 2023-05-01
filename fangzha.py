# 导入所需的库
import openpyxl
from openpyxl.styles import numbers
import tkinter as tk
import tkinter.filedialog as fd
import pandas as pd
import tkinter.messagebox as mb

# 定义第一行的列名
first_row = ["主叫号码", "被叫号码", "呼叫类型", "拨打情况", "开始时间", "拨打时间", "通话时间", "备注", "结束时间", "上一通间隔时间(秒)"]

# 定义一个函数，用于分析excel文件
def analyze_file(filename):
    # 打开excel文件，只读取数据，不读取公式
    wb = openpyxl.load_workbook(filename, data_only=True)
    # 获取活动的工作表
    ws = wb.active
    # 在第一行插入列名
    ws.insert_rows(1)
    for col in range(1, 11):
        ws.cell(row=1, column=col).value = first_row[col-1]
    # 获取最大的行数
    max_row = ws.max_row
    # 遍历每一行，计算结束时间和上一通间隔时间，并设置单元格格式
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=9).value = "=DATE(YEAR(E{}),MONTH(E{}),DAY(E{}))+TIME(HOUR(E{}),MINUTE(E{}),SECOND(E{})+F{})".format(row, row, row, row, row, row, row)
        ws.cell(row=row, column=10).value = "=(E{}-I{})*86400".format(row, row-1)
        ws.cell(row=row, column=5).number_format = 'yyyy/m/d h:mm:ss'
        ws.cell(row=row, column=9).number_format = 'yyyy/m/d h:mm:ss'
        ws.cell(row=row, column=10).number_format = numbers.FORMAT_NUMBER
        ws.cell(row=row, column=2).number_format = numbers.FORMAT_NUMBER
    # 保存分析后的文件，文件名在原来的基础上加上_分析
    wb.save(filename[:-5] + "_分析.xlsx")
    # 调用生成文本结果的函数，并将结果返回给UI界面
    return generate_text(filename[:-5] + "_分析.xlsx")

# 定义一个函数，用于生成文本结果，并保存到excel文件中
def generate_text(filename):
    # 读取分析后的excel文件，只读取第一个工作表
    df = pd.read_excel(filename, sheet_name=0)
    # 将开始时间转换为日期时间格式
    df["开始时间"] = pd.to_datetime(df["开始时间"])
    # 按照主叫号码分组，统计每个号码呼出的次数（只计算呼叫类型为OUTBOUND的）
    grouped = df.groupby("主叫号码")["呼叫类型"].apply(lambda x: x[x == "OUTBOUND"].count())
    # 找出呼出次数最多的号码和次数
    max_number = grouped.idxmax()
    max_count = grouped.max()
    # 计算每个号码的呼出时间（小时为单位）
    df["呼出时间"] = df["开始时间"].dt.hour + df["开始时间"].dt.minute / 60 + df["开始时间"].dt.second / 3600
    # 按照主叫号码和呼出时间（向下取整）分组，统计每个小时内呼出的次数（只计算呼叫类型为OUTBOUND的）
    hourly_grouped = df.groupby(["主叫号码", df["呼出时间"].astype(int)])["呼叫类型"].apply(lambda x: x[x == "OUTBOUND"].count())

    # 按照主叫号码和呼出时间（向下取整）分组，计算每个小时内的平均通话间隔（填充缺失值为0）
    hourly_interval = df.groupby(["主叫号码", df["呼出时间"].astype(int)])["上一通间隔时间(秒)"].mean().fillna(0)

    hourly_count = df.groupby("主叫号码")["呼出时间"].apply(lambda x: x[x.notna()].nunique())

    # 定义一个空列表，用于存放异常的号码
    abnormal_numbers = []
    # 遍历每个号码，如果呼出次数大于50且平均通话间隔小于120秒，则认为是异常号码，加入到列表中
    for number in grouped.index:
        if grouped[number] > 50 and df.groupby("主叫号码")["上一通间隔时间(秒)"].mean()[number] < 120:
            abnormal_numbers.append(number)
    # 遍历每个号码和每个小时，如果连续两个小时内呼出次数都大于20，则认为是异常号码，加入到列表中
    for number in hourly_grouped.index.get_level_values(0).unique():
        for hour in range(23):
            if (number, hour) in hourly_grouped.index and (number, hour + 1) in hourly_grouped.index:
                if hourly_grouped[(number, hour)] > 20 and hourly_grouped[(number, hour + 1)] > 20:
                    abnormal_numbers.append(number)
                    break
    # 去除列表中的重复元素
    abnormal_numbers = list(set(abnormal_numbers))
    # 创建一个空的数据框，用于存放文本结果
    output = pd.DataFrame({
        "分析结果:": []
    })
    # 遍历每个号码，生成文本结果，并添加到数据框中
    for number in grouped.index:
        count = grouped[number]
        interval = df.groupby("主叫号码")["上一通间隔时间(秒)"].mean()[number]
        output.loc[len(output)] = "{}共计呼出{}个电话号码，总平均通话间隔{:.2f}秒。".format(number, count, interval)
        # 如果是异常号码，还要加上警告信息，并设置单元格样式为红色背景和白色字体
        if number in abnormal_numbers:
            output.loc[len(output)] = "注意!!!!----{}----是高风险客户，请立即通知警方!!!!!!!!!!!!!!!!!!".format(number)
            mb.showwarning(title="警告",
                           message="注意!!!!----{}----是高风险客户，请立即通知警方!!!!!!!!!!!!!!!!!!".format(number))
        # 遍历每个小时，生成文本结果，并添加到数据框中
        for hour in range(24):
            if (number, hour) in hourly_grouped.index:
                output.loc[len(output)] = "{}在{}点呼出{}个电话，平均通话间隔{:.2f}秒。".format(number, hour,hourly_grouped[(number, hour)],hourly_interval[(number, hour)])
        # 在每个号码之后添加一个空行
        output.loc[len(output)] = ""
    # 以追加的模式打开分析后的excel文件，并将文本结果保存到一个新的工作表中
    with pd.ExcelWriter(filename, mode="a", engine="openpyxl") as writer:
        output.to_excel(writer, sheet_name="文本结果", index=False)
    # 将文本结果转换为字符串，并返回给UI界面
    return output.to_string(index=False)
# 定义一个函数，用于选择文件并调用分析函数
def choose_file():
    # 弹出一个对话框，让用户选择要分析的excel文件（只支持xlsx格式）
    filename = fd.askopenfilename(title="请选择通话记录文件", filetypes=[("Excel文件", "*.xlsx")])
    # 如果用户选择了文件，则调用分析函数，并将结果显示在文本框中
    if filename:
        result = analyze_file(filename)
        text.delete(1.0, tk.END)
        text.insert(tk.END, result)

# 创建一个窗口，设置标题和大小
window = tk.Tk()
window.title("通话记录分析")
window.geometry("600x600")
# 创建一个标签，显示欢迎信息和使用说明
label = tk.Label(window, text="欢迎使用涉诈通话记录分析工具\n闽侯移动\n请导入XLSX文件", font=("Arial", 16))
label.pack()
# 创建一个按钮，点击后可以选择文件并调用分析函数
button = tk.Button(window, text="导入文件", command=choose_file)
button.pack()
# 创建一个文本框，用于显示分析结果或错误信息
text = tk.Text(window)
text.pack()
# 创建一个按钮，点击后可以关闭窗口
button2 = tk.Button(window, text="关闭窗口", command=window.destroy)
button2.pack()
# 进入窗口的主循环
window.mainloop()
